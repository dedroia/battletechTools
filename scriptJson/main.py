#Tool to: extract json data to Excel, edit in Excel, and then push back to json.

import json
import pathlib
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from copy import deepcopy


# crawls through passed directory
# if changing movespeed, passes any json to editMoveSpeed
# elif pulling data, builds a master dictionary of all json files in any subdirectory provided they are not a 'template'
def crawlDirectories(thisPath: Path):
    global bigDict
    d = {}
    item: Path
    for item in thisPath.glob('**/*.json'):
        if operationIndex == "0":
            editMoveSpeed(item)
        if operationIndex == "1":
            if not 'template' in item.name.lower():
                with item.open() as f:
                    data = json.load(f)
                d = recurseThroughJson(data)
                bigDict[d["Description::Id"]] = d

# if using edit movement option, changes all json keys with "Velocity" and not "Radial" by multiplying by k
def editMoveSpeed(thisPath: Path):
    isFound: bool = False
    with thisPath.open() as f:
        data = json.load(f)
    for key in data:
        if "Velocity" in key and "Radial" not in key:
            data[key] = max(round(data[key] * k),1)
            isFound = True
    with thisPath.open(mode = 'w') as f:
        json.dump(data,f,indent=4,separators=(', ', ' : '))
    if not isFound:
        print(thisPath.name + " does not seem to be a valid movement json.")

# recurses through individual json file
# this "flattens" the json keys, so that it's easier to visualize in excel
# returns new flat dictionary
def recurseThroughJson(data, parent=None, thisDict=None):
    if thisDict is None:
        thisDict = {}
    for key in data:
        if parent is None:
            fullKey = key
        else:
            fullKey = parent + "::" + key
        v = data[key]
        if isinstance(v,dict):
            thisDict = recurseThroughJson(v, fullKey, thisDict)
        elif isinstance(v,list):
            if bool(v) and isinstance(v[0],dict):
                thisDict[fullKey] = '[' + json.dumps(v[0],indent=4,separators=(', ', ' : ')) + ']'
            else:
                thisDict[fullKey] = str(v)
        else:
            thisDict[fullKey] = v
    return thisDict

# sets up the headers in the excel file by pasting all main keys in the first row
def pasteHeaders(thisDict):
    thisColumn = 1
    for key in thisDict:
        ws.cell(row=1,column=thisColumn).value = key
        indexDictionary[key] = thisColumn
        thisColumn += 1


# writes all dictionaries in thisDict to Excel file
# if a 'flattened' key doesn't exist in the header row, does not write contents of json file to Excel
def pasteXlData(thisDict):
    global isListBuilt
    currentRow = 2
    individualItem = {}
    for bKey in thisDict:
        individualItem.clear
        individualItem = thisDict[bKey]
        if not isListBuilt:
            pasteHeaders(individualItem)
            isListBuilt = True
        try:
            for key in individualItem:
                ws.cell(currentRow,indexDictionary[key]).value = individualItem[key]
            currentRow += 1
        except KeyError:
            print(key + ' in "' + individualItem['Description::Id'] + '" not found.')
            print('Removing "' + individualItem['Description::Id'] + '" from excel sheet.')
            cell: Cell = None
            for cell in ws[currentRow]:
               cell.value = None

# prepares new dictionary with correct substructure
# takes 'flattened' keys and changes them to necessary subkeys and dictionaries
# sets value of each key to column to which it is mapped
def buildTemplateDictionary():
    extractDictionary = {}
    column = 1
    while column <= ws.max_column:
        thisHeader = ws.cell(1,column).value
        if "::" in thisHeader:
            extractDictionary = addItemsToList(thisHeader, extractDictionary, column)
        else:
            extractDictionary[thisHeader] = column
        column += 1
    return extractDictionary

# helper function for buildTemplateDictionary
# recurses through passed key, splitting it, and adding the sublevels
# ultimately sets bottom layer key to value    
def addItemsToList(s: object, d: dict, value: object):
    if isinstance(s,str):
        l = s.split("::")
    else:
        l = s
    if not l[0] in d:
        if len(l) > 1:
            d[l[0]] = {}
        else:
            d[l[0]] = value
    if len(l) > 1:
        d[l[0]] = addItemsToList(l[1:], d[l[0]], value)
    return d

# loops through worksheet rows, creates dictionary from each row, and writes dictionary to json
def pushData():
    templateDict = buildTemplateDictionary()
    individualDict = {}
    for ii in range(2, ws.max_row + 1):
        individualDict = deepcopy(templateDict)
        individualDict = makeIndividualDict(individualDict,ii)
        writeJson(individualDict)


# creates dictionary from row based on template dictionary
# loops through all keys, and sets value to mapped column number
def makeIndividualDict(indDict,ii):
    for key in indDict:
        if isinstance(indDict[key],dict):
            indDict[key] = makeIndividualDict(indDict[key],ii)
        else:
            v = ws.cell(ii,indDict[key]).value
            if isinstance(v,str):
                if ("[" in v) and ("]" in v):
                    if (("{" in v) and ("}" in v)):
                        v = json.loads(v)
                    else:
                        v = convertStringToList(v)
            elif v is None:
                v = ""
            indDict[key] = v
    return indDict

# converts strings back to lists where necessary
def convertStringToList(s):
    if len(s) == 2:
        l = []
    else:
        l = s[1:len(s)-1].split(", ")
        if is_integer(l[0]):
            l[:] = [int(x) for x in l]
        elif is_float(l[0]):
            l[:] = [float(x) for x in l]
        else:
            l[:] = [x.replace("'","") for x in l]
    return l

def is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def is_integer(s):
    try:
        int(s)
        return True
    except ValueError:
        return False

# writes indDict dictionary to json file
def writeJson(indDict):
    thisFile: Path = directoryDestination / (str(indDict["Description"]["Id"]) + ".json")
    with thisFile.open('w') as f:
        json.dump(indDict,f,indent=4,separators=(', ', ' : '))

# main variables
directorySource: Path = Path.cwd() / "data"
directorySource.mkdir(exist_ok=True)
directoryDestination: Path = Path.cwd() / "export"
directoryDestination.mkdir(exist_ok = True)

bigDict = {}
indexDictionary = {}
columnHeaders = []
isListBuilt = False
wb: Workbook
ws: Worksheet
workbook_filename: str = 'jsonData.xlsx'

try:
    wb = openpyxl.load_workbook(workbook_filename)
    ws = wb['data']
except FileNotFoundError:
    print('Workbook not found, creating new workbook')
    wb = Workbook()
    ws = wb.active
    ws.title = 'data'
    wb.save('jsonData.xlsx')

# program options
options = ['movement','pull data','push data']
for index, operation in enumerate(options):
    print(str(index) + ': ' + operation)

# input
operationIndex: str = input("Pick operation (enter digit):")
print("\n")
print("\n")

directoryToCrawl: Path

if operationIndex == "0":
    k = float(input("Enter a decimal multiplier for movespeed:"))
    directoryToCrawl = directorySource / "movement"
elif operationIndex == "1":
    pathList = [fld for fld in directorySource.iterdir() if fld.is_dir()]
    for index, fld in enumerate(pathList):
        print(str(index) + ': ' + fld.name)
    directoryIndex: str = input("Pick folder (enter digit):")
    directoryToCrawl = pathList[int(directoryIndex)]


# main function
if operationIndex == "0" or operationIndex == "1":
    crawlDirectories(directoryToCrawl)

# if was collecting data, save output
if operationIndex == "1":
    pasteXlData(bigDict)
    wb.save('jsonData.xlsx')

if operationIndex == "2":
    pushData()